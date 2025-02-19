import openpyxl
import requests
import json
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os
from comunes import *
from geotypes import *

def procesar_lote_locationEsri() -> str: #esta es la función principal que llama a las otras
    #VARIABLES COMUNES
    libro = openpyxl.load_workbook('./Lotes a procesar/entrada_locationEsri.xlsx')
    hoja = libro.active
    max_filas = hoja.max_row
    max_columnas = hoja.max_column
    libro_parametros = openpyxl.load_workbook('parametros.xlsx')
    hoja_parametros = libro_parametros.active
    filter_countries = hoja_parametros.cell(row=7, column=2).value
    filter_categories = hoja_parametros.cell(row=6, column=2).value
    print(str(filter_categories))
    apiKey = 'v1.public.eyJqdGkiOiJlODhlYmRlMy0xMGYwLTRkMzgtYThhZS02MWEwNzA4MTM4ZGUifXYO0jOJMrU6Hy41NjvKMtBXkgLvcS625DbIJkJT9Zk3k4kfs-JGkMfNCqI-JpaGCCbcpr6dz7RsFfXtyk0tH_VROBO8w5NqpY6MqRSBYsFBJ6PWN4pQIGry54VsoHxDdQtU6jOG4QwoL_ZpAxK5mxt7jND1_sVlemBMTn7loOgXFntAzryv4cHaT1vCtAHBaUQoyreQItaDkOcrDCKRxzYvKWRHFm9KvRKL8DMGyFUfINQ4Q28P9PaJjpnABjQVHVBcymnZ12zBB-iw9yKJhK1Pj0Sz7TJ03wfQ4b6u0OcNyyG83wbCWl5mSQL-dwHDVyI42E7Oit0bd4aHeEm9sWU.ZWU0ZWIzMTktMWRhNi00Mzg0LTllMzYtNzlmMDU3MjRmYTkx'
    cabecera_archivo = ["awsEsri-IN-id","awsEsri-IN-singleLine","awsEsri-CUSTOM-orden","awsEsri-Label", "awsEsri-Latitud", "awsEsri-Longitud", 
                        "awsEsri-AddressNumber", "awsEsri-Street", "awsEsri-Neighborhood", "awsEsri-Municipality", 
                        "awsEsri-Subregion", "awsEsri-Region", "awsEsri-PostalCode", "awsEsri-Interpolated",
                        "awsEsri-Categories", "awsEsri-Relevance", "awsEsri-CUSTOM-Json"]
    lista_json = [] # En esta variable se almacenan cada uno de las respuestas de Location
    lista_id = [] # En esta lista se almacenan todos los ID y Single Line que se leyeron
    #--------------------------

    resumenInicio("AWS Location con ESRI", max_filas-1, filter_countries)
    
        #este for itera leyendo una fila del archivo original, retornandola como un diccionario y enviandola a Location.
    for fila in range (2, max_filas + 1): #Lee una por una
        diccionario = leer_singleLine(hoja, fila)
        lista_id.append(diccionario)
        jsonSalida = llamar_location(diccionario, filter_countries, filter_categories, apiKey) #Hace el llamado al servicio
        lista_json.append(jsonSalida)
        print("Se han procesado " + str(fila-1) + " de " + str(max_filas-1) + " registros")  
    print(hora_actual() + " - Se han procesado correctamente con Amazon Location Esri " + str(len(lista_json)) + " registros.\nEspere a la generación del archivo de salida por favor")
    nombre_archivo_generado = generar_archivo_location('OUT_LocatESRI - ', cabecera_archivo, lista_json, lista_id)
    
    return(nombre_archivo_generado)
    
def llamar_location(un_diccionario,un_filter_country, un_filter_category, unaApiKey):
    endpoint = 'https://places.geo.us-east-1.amazonaws.com/places/v0/indexes/location.aws.com.demo.places.Esri.PlaceIndex/search/text?'
    parametros = dict(f='json',key= unaApiKey)
    data = {
            #'FilterCategories' : ['AddressType', 'StreetType', 'IntersectionType', 'PointOfInterestType', 'MunicipalityType'],
            'FilterCategories' : ['AddressType', 'StreetType', 'IntersectionType', 'NeighborhoodType', 'MunicipalityType'],                      
            'FilterCountries' : un_filter_country, #El código de país es de 3 letras            
            #'FilterCountries' : ['PER'], #El código de país es de 3 letras
            'Language' : 'es-419', #El lenguaje es español de latinoamérica
            'MaxResults' : 5,
            'Text': un_diccionario['singleLine'] #Se ingresa la dirección
    }
    respuesta = requests.post(url=endpoint, params=parametros, json=data)
    respuesta_en_json =  respuesta.json()
    return respuesta_en_json

def generar_archivo_location(un_nombre, una_cabecera, una_lista_json, una_lista_id) -> str:
    #la ruta se debe expresar así: "./Documents/Python/Pepito2.xlsx"
    libro_nuevo = Workbook()
    hoja = libro_nuevo.active
    #Esto graba la cabecera del archivo
    print(hora_actual() + " - Se está creando el archivo '" + un_nombre + "' en el directorio local ")
    for campo in range (0, len(una_cabecera)):
        hoja.cell(row=1, column=campo+1).value = una_cabecera[campo]
    grabar_linea_location(una_lista_json, una_lista_id, hoja)
    ruta_archivo = "./Lotes procesados/"+agregar_time_stamp(un_nombre)
    libro_nuevo.save(ruta_archivo)
    print(ruta_archivo)
    
    return ruta_archivo

def grabar_linea_location(una_lista_json, una_lista_id, una_hoja_grabando):
    print(hora_actual() + " - Se comienza a volcar los datos en el archivo de salida")
    fila_escribiendo = 2
    for json in range (0, len(una_lista_json)):
        try:
            for candidate in range (0, len(una_lista_json[json]['Results'])):
                una_hoja_grabando.cell(row=fila_escribiendo, column=1).value = una_lista_id[json]['id']
                una_hoja_grabando.cell(row=fila_escribiendo, column=2).value = una_lista_id[json]['singleLine']
                una_hoja_grabando.cell(row=fila_escribiendo, column=3).value = candidate+1
                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=4).value = una_lista_json[json]['Results'][candidate]['Place']['Label']
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=4).value = "Sin dato"
  
                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=5).value = una_lista_json[json]['Results'][candidate]['Place']['Geometry']['Point'][1]
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=5).value = "Sin dato"
    
                
                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=6).value = una_lista_json[json]['Results'][candidate]['Place']['Geometry']['Point'][0]
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=6).value = "Sin dato"
                    
                    
                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=7).value = una_lista_json[json]['Results'][candidate]['Place']['AddressNumber']
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=7).value = "Sin dato"


                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=8).value = una_lista_json[json]['Results'][candidate]['Place']['Street']
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=8).value = "Sin dato"
                
                
                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=9).value = una_lista_json[json]['Results'][candidate]['Place']['Neighborhood']
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=9).value = "Sin dato"
                
                
                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=10).value = una_lista_json[json]['Results'][candidate]['Place']['Municipality']
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=10).value = "Sin dato"

                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=11).value = una_lista_json[json]['Results'][candidate]['Place']['SubRegion']
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=11).value = "Sin dato"

                
                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=12).value = una_lista_json[json]['Results'][candidate]['Place']['Region']
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=12).value = "Sin dato"



                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=13).value = una_lista_json[json]['Results'][candidate]['Place']['PostalCode']
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=13).value = "Sin dato"
                    
                
                try:
                    interpolated = una_lista_json[json]['Results'][candidate]['Place']['Interpolated']
                    una_hoja_grabando.cell(row=fila_escribiendo, column=14).value = interpolated
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=14).value = "Sin dato"

            
                try:
                    category = una_lista_json[json]['Results'][candidate]['Place']['Categories'][0]
                    una_hoja_grabando.cell(row=fila_escribiendo, column=15).value = category
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=15).value = "Sin dato"

                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=16).value = str(una_lista_json[json]['Results'][candidate]['Relevance'])
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=16).value = "Sin dato"
                    
                try:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=17).value = str(una_lista_json[json])
                except:
                    una_hoja_grabando.cell(row=fila_escribiendo, column=17).value = "Sin dato"

                fila_escribiendo = fila_escribiendo + 1
        except(KeyError):
            print("Sin resultados")
    
        print(hora_actual() + " - Se han volcado " + str(fila_escribiendo-1) + " registros en el archivo")  

