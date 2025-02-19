import openpyxl
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os
from comunes import *

def procesar_lote_arcgis(): #esta es la función principal que llama a las otras
    #VARIABLES COMUNES
    libro = openpyxl.load_workbook('./Lotes a procesar/entrada_arcgis.xlsx')
    hoja = libro.active
    max_filas = hoja.max_row
    libroParametros = openpyxl.load_workbook('parametros.xlsx')
    hojaParametros = libroParametros.active
    countryCode = hojaParametros.cell(row=2, column=2).value
    category = hojaParametros.cell(row=3, column=2).value
    cabecera_archivo = ['arcgis-id','arcgis-singleLine','arcgis-ordenCandidato','arcgis-G01','arcgis-address','arcgis-location_x','arcgis-location_y','arcgis-score','arcgis-Status','arcgis-Addr_type','arcgis-Type','arcgis-AddNum', 
    'arcgis-AddNumFrom','arcgis-AddNumTo','arcgis-StPreType','arcgis-StName','arcgis-Nbrhd','arcgis-District','arcgis-City','arcgis-MetroArea','arcgis-Subregion',
    'arcgis-Region','arcgis-RegionAbbr','arcgis-Territory','arcgis-Zone','arcgis-Postal','arcgis-PostalExt','arcgis-x','arcgis-y','arcGis-Coordenadas','arcgis-ExInfo','arcgis-Json']
    lista_json = [] # En esta variable se almacenan cada uno de las respuestas de ArcGis
    lista_id = [] # En esta lista se almacenan todos los ID y Single Line que se leyeron
    #------------------------------

    resumenInicio("ArcGIS", max_filas-1, countryCode, "entrada_arcgis.xlsx")

    #este for itera leyendo una fila del archivo original, retornandola como un diccionario y enviandola a ArcGis.
    for fila in range (2, max_filas + 1):
        diccionario = leer_singleLine(hoja, fila)
        lista_id.append(diccionario)
        jsonSalida = llamar_arcgis(diccionario,countryCode,category)
        lista_json.append(jsonSalida)
        print("Se han procesado " + str(fila-1) + " de " + str(max_filas-1) + " registros")  
    print(hora_actual() + " - Se han procesado correctamente con ArcGis " + str(len(lista_json)) + " registros.\nEspere a la generación del archivo de salida por favor")
    generar_archivo_arcGis('OUT_arcgis', cabecera_archivo, lista_json, lista_id, libro)
    
def llamar_arcgis(un_diccionario,un_countryCode, unaCategory):
    endpoint = 'https://geocode.arcgis.com/arcgis/rest/services/World/GeocodeServer/findAddressCandidates?'
    parametros = dict(f='json', CountryCode=un_countryCode, outFields='*', maxLocations='5', singleLine=un_diccionario['singleLine'], category=unaCategory)
    respuesta = requests.get(url=endpoint, params=parametros)
    respuesta_en_json = respuesta.json()
    return respuesta_en_json

def generar_archivo_arcGis(un_nombre, una_cabecera, una_lista_json, una_lista_id, un_libro):
    #la ruta se debe expresar así: "./Documents/Python/Pepito2.xlsx"
    libroNuevo = Workbook()
    hoja = libroNuevo.active
    #Esto graba la cabecera del archivo
    print(hora_actual() + " - Se está creando el archivo '" + un_nombre + "' en el directorio local ")
    for campo in range (0, len(una_cabecera)):
        hoja.cell(row=1, column=campo+1).value = una_cabecera[campo]
    grabar_linea_arcGis(una_lista_json, una_lista_id, hoja, un_libro)
    libroNuevo.save("./Lotes procesados/"+agregar_time_stamp(un_nombre))
        

def grabar_linea_arcGis(una_lista_json, una_lista_id, una_hoja_grabando, un_libro):
    print(hora_actual() + " - Se comienza a volcar los datos en el archivo de salida")
    fila_escribiendo = 2
    for json in range (0, len(una_lista_json)):
        contador_point_address = 1
        for candidate in range (0, len(una_lista_json[json]['candidates'])):
            una_hoja_grabando.cell(row=fila_escribiendo, column=1).value = una_lista_id[json]['id']
            una_hoja_grabando.cell(row=fila_escribiendo, column=2).value = una_lista_id[json]['singleLine']
            una_hoja_grabando.cell(row=fila_escribiendo, column=3).value = str(candidate+1)
            
            try:
                if (una_lista_json[json]['candidates'][candidate]['attributes']['Addr_type'] == 'PointAddress' and una_lista_json[json]['candidates'][candidate]['score'] > 90 and contador_point_address == 1):
                    contador_point_address = 0 
                    una_hoja_grabando.cell(row=fila_escribiendo, column=4).value = "primer_candidato_pa"
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=4).value = "Sin datos - ver Json salida"

            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=5).value = una_lista_json[json]['candidates'][candidate]['address']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=5).value = "Sin datos - ver Json salida"
                
            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=6).value = una_lista_json[json]['candidates'][candidate]['location']['x']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=6).value = "Sin datos - ver Json salida"

            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=7).value = una_lista_json[json]['candidates'][candidate]['location']['y']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=7).value = "Sin datos - ver Json salida"

            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=8).value = una_lista_json[json]['candidates'][candidate]['score']
            except(KeyError):
                una_hoja_grabando.cell(row=fila_escribiendo, column=8).value = "Sin datos - ver Json salida"

            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=9).value = una_lista_json[json]['candidates'][candidate]['attributes']['Status']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=9).value = "Sin datos - ver Json salida"
                
            try:    
                una_hoja_grabando.cell(row=fila_escribiendo, column=10).value = una_lista_json[json]['candidates'][candidate]['attributes']['Addr_type']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=10).value = "Sin datos - ver Json salida"    
            
            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=11).value = una_lista_json[json]['candidates'][candidate]['attributes']['Type']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=11).value = "Sin datos - ver Json salida"    

            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=12).value = una_lista_json[json]['candidates'][candidate]['attributes']['AddNum']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=12).value = "Sin datos - ver Json salida"

                
            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=13).value = una_lista_json[json]['candidates'][candidate]['attributes']['AddNumFrom']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=13).value = "Sin datos - ver Json salida"
   
                
            try:    
                una_hoja_grabando.cell(row=fila_escribiendo, column=14).value = una_lista_json[json]['candidates'][candidate]['attributes']['AddNumTo']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=14).value = "Sin datos - ver Json salida"

            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=15).value = una_lista_json[json]['candidates'][candidate]['attributes']['StPreType']
            except(KeyError):
                una_hoja_grabando.cell(row=fila_escribiendo, column=15).value = "Sin datos - ver Json salida"   
                
            try:    
                una_hoja_grabando.cell(row=fila_escribiendo, column=16).value = una_lista_json[json]['candidates'][candidate]['attributes']['StName']
            except(KeyError):
                una_hoja_grabando.cell(row=fila_escribiendo, column=16).value = "Sin datos - ver Json salida"

            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=17).value = una_lista_json[json]['candidates'][candidate]['attributes']['Nbrhd']
                
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=17).value = "Sin datos - ver Json salida"

            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=18).value = una_lista_json[json]['candidates'][candidate]['attributes']['District']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=18).value = "Sin datos - ver Json salida"
            
            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=19).value = una_lista_json[json]['candidates'][candidate]['attributes']['City']
                
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=19).value = "Sin datos - ver Json salida"

            try:    
                una_hoja_grabando.cell(row=fila_escribiendo, column=20).value = una_lista_json[json]['candidates'][candidate]['attributes']['MetroArea']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=20).value = "Sin datos - ver Json salida"    
                
            try:    
                una_hoja_grabando.cell(row=fila_escribiendo, column=21).value = una_lista_json[json]['candidates'][candidate]['attributes']['Subregion']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=21).value = "Sin datos - ver Json salida"

            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=22).value = una_lista_json[json]['candidates'][candidate]['attributes']['Region']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=22).value = "Sin datos - ver Json salida"    
                
            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=23).value = una_lista_json[json]['candidates'][candidate]['attributes']['RegionAbbr']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=23).value = "Sin datos - ver Json salida"    
            
            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=24).value = una_lista_json[json]['candidates'][candidate]['attributes']['Territory']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=24).value = "Sin datos - ver Json salida"    
            
            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=25).value = una_lista_json[json]['candidates'][candidate]['attributes']['Zone']            
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=25).value = "Sin datos - ver Json salida"

            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=26).value = una_lista_json[json]['candidates'][candidate]['attributes']['Postal']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=26).value = "Sin datos - ver Json salida"

            try:        
                una_hoja_grabando.cell(row=fila_escribiendo, column=27).value = una_lista_json[json]['candidates'][candidate]['attributes']['PostalExt']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=27).value = "Sin datos - ver Json salida"    
            
            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=28).value = una_lista_json[json]['candidates'][candidate]['attributes']['X']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=28).value = "Sin datos - ver Json salida"

            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=29).value = una_lista_json[json]['candidates'][candidate]['attributes']['Y']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=29).value = "Sin datos - ver Json salida"
                
            try:    
                una_hoja_grabando.cell(row=fila_escribiendo, column=30).value = concatenarCoordenadas(str(una_lista_json[json]['candidates']\
                [candidate]['attributes']['Y']), str(una_lista_json[json]['candidates'][candidate]['attributes']['X']))
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=30).value = "Sin datos - ver Json salida"    
                
            try:    
                una_hoja_grabando.cell(row=fila_escribiendo, column=31).value = una_lista_json[json]['candidates'][candidate]['attributes']['ExInfo']
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=31).value = "Sin datos - ver Json salida"
            
            try:
                una_hoja_grabando.cell(row=fila_escribiendo, column=32).value = str(una_lista_json[json]['candidates'][candidate])
            except:
                una_hoja_grabando.cell(row=fila_escribiendo, column=32).value = "Sin datos - ver Json salida"

            fila_escribiendo = fila_escribiendo + 1
        #un_libro.save("./Lotes procesados/temp.xlsx")
        contador_point_address = 1
    print(hora_actual() + " - Se han volcado " + str(fila_escribiendo-1) + " registros en el archivo")   