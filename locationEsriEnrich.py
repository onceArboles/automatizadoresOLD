import openpyxl
import requests
import json
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os
from comunes import *
from geotypes import *
from estadosMotivos import *

def procesar_lote_locationEsriEnrich() -> str: #esta es la función principal que llama a las otras
    #VARIABLES COMUNES
    libro = openpyxl.load_workbook('./Lotes a procesar/entrada_locationEsri.xlsx')
    hoja = libro.active
    max_filas = hoja.max_row
    max_columnas = hoja.max_column
    libro_parametros = openpyxl.load_workbook('parametros.xlsx')
    hoja_parametros = libro_parametros.active
    filter_countries = hoja_parametros.cell(row=7, column=2).value
    filter_categories = hoja_parametros.cell(row=6, column=2).value
    
    apiKey = 'v1.public.eyJqdGkiOiJlODhlYmRlMy0xMGYwLTRkMzgtYThhZS02MWEwNzA4MTM4ZGUifXYO0jOJMrU6Hy41NjvKMtBXkgLvcS625DbIJkJT9Zk3k4kfs-JGkMfNCqI-JpaGCCbcpr6dz7RsFfXtyk0tH_VROBO8w5NqpY6MqRSBYsFBJ6PWN4pQIGry54VsoHxDdQtU6jOG4QwoL_ZpAxK5mxt7jND1_sVlemBMTn7loOgXFntAzryv4cHaT1vCtAHBaUQoyreQItaDkOcrDCKRxzYvKWRHFm9KvRKL8DMGyFUfINQ4Q28P9PaJjpnABjQVHVBcymnZ12zBB-iw9yKJhK1Pj0Sz7TJ03wfQ4b6u0OcNyyG83wbCWl5mSQL-dwHDVyI42E7Oit0bd4aHeEm9sWU.ZWU0ZWIzMTktMWRhNi00Mzg0LTllMzYtNzlmMDU3MjRmYTkx'
    cabecera_location = ["awsEsri-IN-id","awsEsri-IN-singleLine","awsEsri-CUSTOM-orden","awsEsri-Label", "awsEsri-Latitud", 
                         "awsEsri-Longitud", "awsEsri-AddressNumber", "awsEsri-Street", "awsEsri-Neighborhood", "awsEsri-Municipality", 
                         "awsEsri-Subregion", "awsEsri-Region", "awsEsri-PostalCode", "awsEsri-Interpolated","awsEsri-Categories", 
                         "awsEsri-Relevance", "awsEsri-CUST-TipoGeo", "awsEsri-CUST-Estado", "awsEsri-CUST-Motivo", "awsEsri-CUSTOM-Json"]
    
    cabecera_address = ["IDCliente","Estado", "Motivo", "TipoGeo","Calle","Altura", "L2-Departamento", "L3-Provincia", "L4-Distrito", "L5-Barrio", "CódigoPostal","Latitud", "Longitud"]
    lista_json = [] # En esta variable se almacenan cada uno de las respuestas de Location
    lista_id = [] # En esta lista se almacenan todos los ID y Single Line que se leyeron
    #--------------------------

    hora_inicio = resumenInicio("AWS Location con ESRI y Normalización", max_filas-1, filter_countries, "entrada_locationEsri.xlsx")
    
        #este for itera leyendo una fila del archivo original, retornandola como un diccionario y enviandola a Location.
    for fila in range (2, max_filas + 1): #Lee una por una
        diccionario = leer_singleLine(hoja, fila)
        lista_id.append(diccionario)
        jsonSalida = llamar_location(diccionario, filter_countries, filter_categories, apiKey) #Hace el llamado al servicio
        lista_json.append(jsonSalida)
        print("Se han procesado " + str(fila-1) + " de " + str(max_filas-1) + " registros")  
    print(hora_actual() + " - Se han procesado correctamente con Amazon Location " + str(len(lista_json)) + " registros.\nEspere a la generación del archivo de salida por favor")
    nombre_archivo_generado = generar_archivo_location('OUT_LocatESRI_Enrich - ', cabecera_location, cabecera_address, lista_json, lista_id)
    #print("Inicio del proceso a las  ")
    return(nombre_archivo_generado)
    
def llamar_location(un_diccionario,un_filter_country, un_filter_category, unaApiKey):
    endpoint = 'https://places.geo.us-east-1.amazonaws.com/places/v0/indexes/location.aws.com.demo.places.Esri.PlaceIndex/search/text?'
    parametros = dict(f='json',key= unaApiKey)
    data = {
            #'FilterCategories' : ['AddressType', 'StreetType', 'IntersectionType', 'PointOfInterestType', 'MunicipalityType'],
            'FilterCategories' : ['AddressType', 'StreetType', 'IntersectionType', 'NeighborhoodType', 'MunicipalityType'],
            'FilterCountries' : un_filter_country, #El código de país es de 3 letras            
            #'FilterCountries' : ['PER'], #El código de país es de 3 letras
            'Language' : 'es-419', #El lenguaje es español de latinoamérica
            'MaxResults' : 5,
            'Text': un_diccionario['singleLine'] #Se ingresa la dirección
    }
    respuesta = requests.post(url=endpoint, params=parametros, json=data)
    respuesta_en_json =  respuesta.json()
    return respuesta_en_json

def generar_archivo_location(un_nombre, una_cabecera1,una_cabecera2, una_lista_json, una_lista_id) -> str:
    #la ruta se debe expresar así: "./Documents/Python/Pepito2.xlsx"
    libro_nuevo = Workbook()
    hoja1 = libro_nuevo.active
    hoja1.title = "SalidaLocat"
    hoja2 = libro_nuevo.create_sheet("LocatMejorCand")
    hoja3 = libro_nuevo.create_sheet("FinalCliente")
    print(hora_actual() + " - Se está creando el archivo '" + un_nombre + "' en el directorio local ")
    #Esto graba la cabecera del archivo
    for campo in range (0, len(una_cabecera1)):
        hoja1.cell(row=1, column=campo+1).value = una_cabecera1[campo]
        hoja2.cell(row=1, column=campo+1).value = una_cabecera1[campo]
    for campo in range (0, len(una_cabecera2)):
        hoja3.cell(row=1, column=campo+1).value = una_cabecera2[campo]

    grabar_linea_location(una_lista_json, una_lista_id, hoja1, hoja2, hoja3)
    ruta_archivo = "./Lotes procesados/"+agregar_time_stamp(un_nombre)
    libro_nuevo.save(ruta_archivo)

    return ruta_archivo #Hoy no se usa

def grabar_linea_location(una_lista_json, una_lista_id, una_hoja1_grabando, una_hoja2_grabando, una_hoja3_grabando):
    print(hora_actual() + " - Se comienza a volcar los datos en el archivo de salida")
    fila_hoja1_escribiendo = 2
    fila_hoja23_escribiendo = 2
    for json in range (0, len(una_lista_json)):
        #aca traigo el diccionario normalizado y el mejor candidato
            if (len(una_lista_json[json]["Results"]) > 0):
                candidatos_seleccionados = trabajar_json(una_lista_json[json], una_lista_id[json])
                mejor_candidato = candidatos_seleccionados[1]
                for candidato in range (0, len(candidatos_seleccionados[0])):
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=1).value = candidatos_seleccionados[0][candidato]['id']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=2).value = candidatos_seleccionados[0][candidato]['singleLine']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=3).value = candidatos_seleccionados[0][candidato]['orden']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=4).value = candidatos_seleccionados[0][candidato]['label']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=5).value = candidatos_seleccionados[0][candidato]['lat']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=6).value = candidatos_seleccionados[0][candidato]['lon']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=7).value = candidatos_seleccionados[0][candidato]['addressNumber']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=8).value = candidatos_seleccionados[0][candidato]['street']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=9).value = candidatos_seleccionados[0][candidato]['neighborhood']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=10).value = candidatos_seleccionados[0][candidato]['municipality']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=11).value = candidatos_seleccionados[0][candidato]['subregion']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=12).value = candidatos_seleccionados[0][candidato]['region']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=13).value = candidatos_seleccionados[0][candidato]['postalCode']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=14).value = candidatos_seleccionados[0][candidato]['interpolated']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=15).value = candidatos_seleccionados[0][candidato]['category']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=16).value = candidatos_seleccionados[0][candidato]['relevance']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=17).value = candidatos_seleccionados[0][candidato]['tipoGeo'][0]
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=18).value = candidatos_seleccionados[0][candidato]['estado']
                    una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=20).value = candidatos_seleccionados[0][candidato]['json']
                    fila_hoja1_escribiendo = fila_hoja1_escribiendo + 1
                    print("Se han volcado " + str(fila_hoja1_escribiendo) + " registros en la hoja 'SalidaLocat")
                    
                
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=1).value = candidatos_seleccionados[0][mejor_candidato]['id']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=2).value = candidatos_seleccionados[0][mejor_candidato]['singleLine']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=3).value = candidatos_seleccionados[0][mejor_candidato]['orden']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=4).value = candidatos_seleccionados[0][mejor_candidato]['label']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=5).value = candidatos_seleccionados[0][mejor_candidato]['lat']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=6).value = candidatos_seleccionados[0][mejor_candidato]['lon']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=7).value = candidatos_seleccionados[0][mejor_candidato]['addressNumber']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=8).value = candidatos_seleccionados[0][mejor_candidato]['street']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=9).value = candidatos_seleccionados[0][mejor_candidato]['neighborhood']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=10).value = candidatos_seleccionados[0][mejor_candidato]['municipality']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=11).value = candidatos_seleccionados[0][mejor_candidato]['subregion']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=12).value = candidatos_seleccionados[0][mejor_candidato]['region']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=13).value = candidatos_seleccionados[0][mejor_candidato]['postalCode']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=14).value = candidatos_seleccionados[0][mejor_candidato]['interpolated']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=15).value = candidatos_seleccionados[0][mejor_candidato]['category']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=16).value = candidatos_seleccionados[0][mejor_candidato]['relevance']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=17).value = candidatos_seleccionados[0][mejor_candidato]['tipoGeo'][0]
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=18).value = candidatos_seleccionados[0][mejor_candidato]['estado']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=20).value = candidatos_seleccionados[0][mejor_candidato]['json']
                print("Se han volcado " + str(fila_hoja23_escribiendo) + " registros en la hoja 'LocatMejorCand'")
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=1).value = candidatos_seleccionados[0][mejor_candidato]['id']
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=2).value = candidatos_seleccionados[0][mejor_candidato]['estado']
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=4).value = candidatos_seleccionados[0][mejor_candidato]['tipoGeo'][0]
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=5).value = candidatos_seleccionados[0][mejor_candidato]['street']
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=6).value = candidatos_seleccionados[0][mejor_candidato]['addressNumber']
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=7).value = candidatos_seleccionados[0][mejor_candidato]['region']
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=8).value = candidatos_seleccionados[0][mejor_candidato]['subregion']
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=9).value = candidatos_seleccionados[0][mejor_candidato]['municipality']
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=10).value = candidatos_seleccionados[0][mejor_candidato]['neighborhood']
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=11).value = candidatos_seleccionados[0][mejor_candidato]['postalCode']
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=12).value = candidatos_seleccionados[0][mejor_candidato]['lat']
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=13).value = candidatos_seleccionados[0][mejor_candidato]['lon']
                print("Se han volcado " + str(fila_hoja23_escribiendo) + " registros en la hoja 'FinalCliente'")
                fila_hoja23_escribiendo = fila_hoja23_escribiendo + 1

            else:
                una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=1).value = una_lista_id[json]['id']
                una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=2).value = una_lista_id[json]['singleLine']
                una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=18).value = "NE"            
                una_hoja1_grabando.cell(row=fila_hoja1_escribiendo, column=20).value = str(una_lista_json[json])
                
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=1).value = una_lista_id[json]['id']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=2).value = una_lista_id[json]['singleLine']
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=18).value = "NE"            
                una_hoja2_grabando.cell(row=fila_hoja23_escribiendo, column=20).value = str(una_lista_json[json])
                
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=1).value = una_lista_id[json]['id']
                una_hoja3_grabando.cell(row=fila_hoja23_escribiendo, column=2).value = "NE"                      
                
                fila_hoja1_escribiendo = fila_hoja23_escribiendo + 1
                fila_hoja23_escribiendo = fila_hoja23_escribiendo + 1
                print("Zero results - No hay candidatos de Location")
            

def trabajar_json(un_json_location, un_json_entrada):
    lista_candidatos = []
    mejor_candidato = [0, ["",99], "-"]
    #en el arreglo de arriba "0" es órden del candidato, 99 un tipo geo seteado para que se 
    # tome si o si el primer candidao antes y "-" es el estado
    for candidate in range (0, len(un_json_location['Results'])):
        diccionario_candidato = {}

        diccionario_candidato['id'] = un_json_entrada['id']
        diccionario_candidato['singleLine'] = un_json_entrada['singleLine']
        diccionario_candidato['orden'] = candidate+1

        try:
            diccionario_candidato['label'] = un_json_location['Results'][candidate]['Place']['Label']
        except:
            diccionario_candidato['label'] = "s/d"
                
        try:
            diccionario_candidato['lat'] = un_json_location['Results'][candidate]['Place']['Geometry']['Point'][1]
        except:
            diccionario_candidato['lat'] = "s/d"    
                
        try:
            diccionario_candidato['lon'] = un_json_location['Results'][candidate]['Place']['Geometry']['Point'][0]
        except:
            diccionario_candidato['lon'] = "s/d"                    
                    
        try:
            diccionario_candidato['addressNumber'] = un_json_location['Results'][candidate]['Place']['AddressNumber']
        except:
            diccionario_candidato['addressNumber'] = "s/d"

        try:
            diccionario_candidato['street'] = un_json_location['Results'][candidate]['Place']['Street']
        except:
            diccionario_candidato['street'] = "s/d"                
                
        try:
            diccionario_candidato['neighborhood'] = un_json_location['Results'][candidate]['Place']['Neighborhood']
        except:
            diccionario_candidato['neighborhood'] = "s/d"                
                
        try:
            diccionario_candidato['municipality'] = un_json_location['Results'][candidate]['Place']['Municipality']
        except:
            diccionario_candidato['municipality'] = "s/d"

        try:
            diccionario_candidato['subregion'] = un_json_location['Results'][candidate]['Place']['SubRegion']
        except:
            diccionario_candidato['subregion'] = "s/d"
               
        try:
            diccionario_candidato['region'] = un_json_location['Results'][candidate]['Place']['Region']
        except:
            diccionario_candidato['region'] = "s/d"

        try:
            diccionario_candidato['postalCode'] = un_json_location['Results'][candidate]['Place']['PostalCode']
        except:
            diccionario_candidato['postalCode'] = "s/d"
                            
        try:
            interpolated = bool(un_json_location['Results'][candidate]['Place']['Interpolated'])
            diccionario_candidato['interpolated'] = un_json_location['Results'][candidate]['Place']['Interpolated']
        except:
            diccionario_candidato['interpolated'] = "s/d"
            
        try:
            category = un_json_location['Results'][candidate]['Place']['Categories'][0]
            diccionario_candidato['category'] = category
        except:
            diccionario_candidato['category'] = "s/d"

        try:
            relevance = un_json_location['Results'][candidate]['Relevance']
            diccionario_candidato['relevance'] = str(relevance)
        except:
            diccionario_candidato['relevance'] = "s/d"
        try:        
            tipo_geo = location_esri_PE(category, relevance, interpolated)
            diccionario_candidato['tipoGeo'] = tipo_geo
        except:
            diccionario_candidato['tipoGeo'] = "Error"
        
        try:
            estado = estado_openpyxl_basico(tipo_geo[0])
            diccionario_candidato['estado'] = estado
        except:
            diccionario_candidato['estado'] = "Error"
            
        try:
            #CALCULO DEL CANDIDATO A DEVOLVER
            if (tipo_geo[1] < mejor_candidato[1][1]):
                mejor_candidato[0] = candidate
                mejor_candidato[1] = tipo_geo
                mejor_candidato[2] = estado
        except:
                mejor_candidato[0] = candidate
                mejor_candidato[1] = ["-",99]
                mejor_candidato[2] = "s/d"
            
        #--------------------------------------------                    
        try:
            diccionario_candidato['json'] = str(un_json_location)
        except:
            diccionario_candidato['json'] = "-"
        
        lista_candidatos.append(diccionario_candidato)
    return [lista_candidatos, mejor_candidato[0]]