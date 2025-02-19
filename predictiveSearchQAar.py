import openpyxl
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os
from tokenSecurityQA import *
from comunes import *

def procesar_lote_predictiveAR():
    #VARIABLES COMUNES
    token = pedir_token()
    libro = openpyxl.load_workbook('./Lotes a procesar/entrada_predictiveQA.xlsx')
    hoja = libro.active
    max_filas = hoja.max_row
    
    libroParametros = openpyxl.load_workbook('parametros.xlsx')
    hojaParametros = libroParametros.active
    
    #en cabeecera_archivo se definen los campos de la cabecera del archivo de salida
    cabecera_archivo = ['predAR-id','predAR-#order','predAR-singleLine','predAR-street','predAR-houseNumber','predAR-postalCode','predAR-latitude','predAR-longitude','predAR-label','predAR-locationType','predAR-suggestedItem','predAR-maps','predAR-geoType','predAR-level1','predAR-level2','predAR-level3','predAR-level4','predAR-level5','predAR-Json']
    lista_json = [] # En esta variable se almacenan cada uno de las respuestas de Búsqueda Predictiva
    lista_id = [] # En esta lista se almacenan todos los ID y Single Line que se leyeron
    #--------------------------------

    resumenInicio("Predictive", max_filas-1, "AR", "entrada_predictiveQA.xlsx")

    #este for itera leyendo una fila del archivo original, retornandola como un diccionario y enviandola a ArcGis.
    for fila in range (2, max_filas + 1):
        diccionario = leer_singleLine(hoja, fila)
        lista_id.append(diccionario)
        jsonSalida = llamar_predictiveAR(diccionario,token)
        lista_json.append(jsonSalida)
        print("Se han procesado " + str(fila-1) + " de " + str(max_filas-1) + " registros")  
    print(hora_actual() + " - Se han procesado correctamente con Búsqueda Predictiva " + str(len(lista_json)) + " registros.\nEspere a la generación del archivo de salida por favor")
    generar_archivo_predictiveAR('Lote_procesado_predictiveAR - ', cabecera_archivo, lista_json, lista_id)
    input("Presione cualquier tecla para finalizar")


def llamar_predictiveAR(un_diccionario, un_token):
    params = {
    'stringSearch': un_diccionario['singleLine'],
    'level1': 'AR'}
    
    headers = {
    'Authorization': un_token,
    'Content-Type': 'application/json'}
    
    endpoint = 'https://wsqa.merlindataquality.com/predictivesearch/search?'
    
    response = requests.get(url=endpoint, params=params, headers=headers)
    respuestaJson = response.json()
    return(respuestaJson)
    
def generar_archivo_predictiveAR(un_nombre, una_cabecera, una_lista_json, una_lista_id):
    #la ruta se debe expresar así: "./Documents/Python/Pepito2.xlsx"
    libroNuevo = Workbook()
    hoja = libroNuevo.active
    #Esto graba la cabecera del archivo
    print(hora_actual() + " - Se está creando el archivo '" + un_nombre + "' en el directorio local ")
    for campo in range (0, len(una_cabecera)):
        hoja.cell(row=1, column=campo+1).value = una_cabecera[campo]
    grabar_linea_predictiveAR(una_lista_json, una_lista_id, hoja)
    libroNuevo.save("./Lotes procesados/"+agregar_time_stamp(un_nombre))
    
def grabar_linea_predictiveAR(una_lista_json, una_lista_id, una_hoja_grabando):
    print(hora_actual() + " - Se comienza a volcar los datos en el archivo de salida")
    fila_escribiendo = 2
    for json in range (0, len(una_lista_json)):
        for candidate in range (0, len(una_lista_json[json]['mSuggest'])):
            una_hoja_grabando.cell(row=fila_escribiendo, column=1).value = una_lista_id[json]['id']
            una_hoja_grabando.cell(row=fila_escribiendo, column=2).value = candidate+1            
            una_hoja_grabando.cell(row=fila_escribiendo, column=3).value = una_lista_id[json]['singleLine']
            una_hoja_grabando.cell(row=fila_escribiendo, column=4).value = una_lista_json[json]['mSuggest'][candidate]['street']
            una_hoja_grabando.cell(row=fila_escribiendo, column=5).value = una_lista_json[json]['mSuggest'][candidate]['houseNumber']
            una_hoja_grabando.cell(row=fila_escribiendo, column=6).value = una_lista_json[json]['mSuggest'][candidate]['postalCode']
            una_hoja_grabando.cell(row=fila_escribiendo, column=7).value = una_lista_json[json]['mSuggest'][candidate]['latitude']
            una_hoja_grabando.cell(row=fila_escribiendo, column=8).value = una_lista_json[json]['mSuggest'][candidate]['longitude']
            una_hoja_grabando.cell(row=fila_escribiendo, column=9).value = una_lista_json[json]['mSuggest'][candidate]['label']
            una_hoja_grabando.cell(row=fila_escribiendo, column=10).value = una_lista_json[json]['mSuggest'][candidate]['locationType']
            una_hoja_grabando.cell(row=fila_escribiendo, column=11).value = una_lista_json[json]['mSuggest'][candidate]['suggestedItem']
            una_hoja_grabando.cell(row=fila_escribiendo, column=12).value = una_lista_json[json]['mSuggest'][candidate]['maps']
            una_hoja_grabando.cell(row=fila_escribiendo, column=13).value = una_lista_json[json]['mSuggest'][candidate]['geoType']
            una_hoja_grabando.cell(row=fila_escribiendo, column=14).value = una_lista_json[json]['mSuggest'][candidate]['level1']
            una_hoja_grabando.cell(row=fila_escribiendo, column=15).value = una_lista_json[json]['mSuggest'][candidate]['level2']
            una_hoja_grabando.cell(row=fila_escribiendo, column=16).value = una_lista_json[json]['mSuggest'][candidate]['level3']
            una_hoja_grabando.cell(row=fila_escribiendo, column=17).value = una_lista_json[json]['mSuggest'][candidate]['level4']
            una_hoja_grabando.cell(row=fila_escribiendo, column=18).value = una_lista_json[json]['mSuggest'][candidate]['level5']
            una_hoja_grabando.cell(row=fila_escribiendo, column=19).value = str(una_lista_json[json]['mSuggest'][candidate])
            
            fila_escribiendo = fila_escribiendo + 1
    print(hora_actual() + " - Se han volcado " + str(fila_escribiendo-2) + " registros en el archivo")  
    

