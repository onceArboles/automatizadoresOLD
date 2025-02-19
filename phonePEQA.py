import requests
from tokenSecurityQA import *
import openpyxl
import requests
import json
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os
from comunes import agregar_time_stamp
from comunes import hora_actual

def procesarLotePhonePeQA () -> None:
    """Esta funcion procesa un lote de teléfonos de Perú en QAs. A partir de un archivo xlsx de teléfonos con la estructura definida, 
    retorna un archivo de mismo formato con la salida formateada y tabulada.
    Endpoint: https://wsqa.merlindataquality.com/phone/normalize
    """

    #VARIABLES COMUNES
    token = pedir_token()
    libro = openpyxl.load_workbook('./Lotes a procesar/entrada_phonePE.xlsx')
    hoja = libro.active
    max_filas = hoja.max_row
    
    cabecera_archivo = ['PhonePEQAs_id_register', 'PhonePEQAs_status','PhonePEQAs_statusReason','PhonePEQAs_level1', 
                        'PhonePEQAs_level2', 'PhonePEQAs_level3', 'PhonePEQAs_level4', 'PhonePEQAs_level5','PhonePEQAs_postalCode', 
                        'PhonePEQAs_ddi', 'PhonePEQAs_ddn','PhonePEQAs_characteristic','PhonePEQAs_phoneNumber', 
                        'PhonePEQAs_validated','PhonePEQAs_additionalData', 'PhonePEQAs_fullCellPhoneNumber','PhonePEQAs_fullPhone',
                        'PhonePEQAs_directory', 'PhonePEQAs_doNotCallRegistry', 'PhonePEQAs_postalCode', 'IN_level1','IN_level2', 
                        'IN_level3', 'IN_level4', 'IN_level5', 'IN_phoneNumber', 'IN_characteristic', 'IN_postalCode', 'IN_prefix',
                        'IN_additionalData','JsonCompleto'
                        ]

        #En esta variable se almacenan los ID y las direcciones que se leen
    lista_id = [] 
    
    #En esta variable se almacenan cada uno de las respuestas de la API
    lista_json = []
    
    #----------------------------

    resumenInicio("Phone", max_filas-1, "Pe", "entrada_phonePE.xlsx")
    
    for fila in range (2, max_filas + 1):
        diccionario = leer_phone_peru(hoja, fila) #leo una fila del archivo
        lista_id.append(diccionario) #agrego la fila leida a la lista de direcciones leidas        
        jsonSalida = llamar_PhonePe(diccionario, pedir_token()) #llamo a Phone
        lista_json.append(jsonSalida) #agrego la salida de phone a la lista de Json de salida
        print("Se han procesado " + str(fila-1) + " de " + str(max_filas-1) + " registros con Phone Perú en QAS")  
    print(hora_actual() + " - Se han procesado correctamente con Phone Perú en Producción " + str(len(lista_json)) + " registros.\nEspere a la generación del archivo de salida por favor")
    generar_archivo_phonePe('Lote_procesado_phonePEprod - ', cabecera_archivo, lista_json, lista_id) 
    
    
def leer_phone_peru(una_hoja: openpyxl, una_fila: int) -> dict:
    """La función lee una fila del archivo de entrada y la almacena en un diccionario que se retrorna a la función principal

    Args:
        una_hoja (openpyxl): Una hoja desde la cual leer los valores
        una_fila (int): número de fila que se está leyendo

    Returns:
        dict: Diccionario con la fila leída en el archivo.
    """
    
    diccionario = {}
    diccionario['idRegister'] = una_hoja.cell(row = una_fila, column = 1).value
    
    if una_hoja.cell(row = una_fila, column = 2).value == None:
        diccionario['level2'] = ''
    else:
        diccionario['level2'] = una_hoja.cell(row = una_fila, column = 2).value


    if una_hoja.cell(row = una_fila, column = 3).value == None:
        diccionario['level3'] = ''
    else:
        diccionario['level3'] = una_hoja.cell(row = una_fila, column = 3).value
        
        
    if una_hoja.cell(row = una_fila, column = 4).value == None:
        diccionario['level4'] = ''
    else:
        diccionario['level4'] = una_hoja.cell(row = una_fila, column = 4).value  
        
        
    if una_hoja.cell(row = una_fila, column = 5).value == None:
        diccionario['level5'] = ''
    else:
        diccionario['level5'] = una_hoja.cell(row = una_fila, column = 5).value
        

    if una_hoja.cell(row = una_fila, column = 6).value == None:
        diccionario['phoneNumber'] = ''
    else:
        diccionario['phoneNumber'] = una_hoja.cell(row = una_fila, column = 6).value

    if una_hoja.cell(row = una_fila, column = 7).value == None:
        diccionario['characteristic'] = ''
    else:
        diccionario['characteristic'] = una_hoja.cell(row = una_fila, column = 7).value

    if una_hoja.cell(row = una_fila, column = 8).value == None:
        diccionario['postalCode'] = ''
    else:
        diccionario['postalCode'] = una_hoja.cell(row = una_fila, column = 8).value

    if una_hoja.cell(row = una_fila, column = 9).value == None:
        diccionario['prefix'] = ''
    else:
        diccionario['prefix'] = una_hoja.cell(row = una_fila, column = 9).value
        
    if una_hoja.cell(row = una_fila, column = 10).value == None:
        diccionario['additionalData'] = ''
    else:
        diccionario['additionalData'] = una_hoja.cell(row = una_fila, column = 10).value

    return diccionario
    
    
def llamar_PhonePe(un_diccionario: dict, un_token: str) -> json:
    """La función toma un diccionario con un único json de entrada y un token de QAs y lo envía a Phone Peru. 
        Retorna un Json con la salida.
        
        Endpoint = "https://wsqa.merlindataquality.com/phone/normalize"

    Args:
        un_diccionario (dict): diccionario que contiene el teléfono tal como lo espera la mensajería
        un_token (str): un token del ambiente de QAs

    Returns:
        json: Devuelve un json con el teléfono normalizado por Merlín.
    """
    
    url = "https://wsqa.merlindataquality.com/phone/normalize"

    payload = json.dumps({
        "level1": "PE",
        "level2": un_diccionario['level2'],
        "level3": un_diccionario['level3'],
        "level4": un_diccionario['level4'],
        "level5": un_diccionario['level5'],
        "phoneNumber": un_diccionario['phoneNumber'],
        "characteristic": un_diccionario['characteristic'],
        "postalCode": un_diccionario['postalCode'],
        "prefix": un_diccionario['prefix'],
        "additionalData": un_diccionario['additionalData']
    })
    headers = {
        'Content-Type': 'application/json',
        'Authorization': un_token
    }
    response = requests.request("POST", url, headers=headers, data=payload)
    
    respuesta_json = response.json()
    return respuesta_json


def generar_archivo_phonePe(un_nombre: str, una_cabecera: list, una_lista_json: list, una_lista_id: list) -> None:
    """La función genera una archivo xlxs de salida con todas los jsons devueltos por Phone Peru, tabulados. 
    Incluye también la petición para cada salida.

    Args:
        un_nombre (str): Nombre del archivo a crear.
        una_cabecera (list): Una lista con los nombres de los campos para generar el archivo de salida.
        una_lista_json (list): Una lista de diccionarios/json con las respuestas devueltas por PhonePe
        una_lista_id (list): Una lista de diccionarios/json con las filas leidas en el archivo de entrada.
    """
    #la ruta se debe expresar así: "./Documents/Python/Pepito2.xlsx"
    libroNuevo = Workbook()
    hoja = libroNuevo.active
    #Esto graba la cabecera del archivo
    print(hora_actual() + " - Se está creando el archivo '" + un_nombre + "' en el directorio local ")
    for campo in range (0, len(una_cabecera)):
        hoja.cell(row=1, column=campo+1).value = una_cabecera[campo]
    grabar_linea_phonePe(una_lista_json, una_lista_id, hoja)
    libroNuevo.save("./Lotes procesados/"+agregar_time_stamp(un_nombre))


def grabar_linea_phonePe(una_lista_json: list, una_lista_id: list, una_hoja_grabando: openpyxl) -> None:
    """Graba cada uno de los jsons/diccionarios devueltos por PhonePe en el archivo de salida

    Args:
        una_lista_json (list): Una lista de diccionarios/json con las respuestas devueltas por PhonePe
        una_lista_id (list):  Una lista de diccionarios/json con las filas leidas en el archivo de entrada.
        una_hoja_grabando (openpyxl): Una hoja en la cual grabar los registros
    """
    print(hora_actual() + " - Se comienza a volcar los datos en el archivo de salida")
    fila_escribiendo = 2
    for json in range (0, len(una_lista_id)):
        una_hoja_grabando.cell(row=fila_escribiendo, column=1).value = una_lista_id[json]['idRegister']
        
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=2).value = una_lista_json[json]['status']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=2).value = 'sin datos'

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=3).value = una_lista_json[json]['statusReason']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=3).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=4).value = una_lista_json[json]['nPhone']['level1']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=4).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=5).value = una_lista_json[json]['nPhone']['level2']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=5).value = 'sin datos'            

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=6).value = una_lista_json[json]['nPhone']['level3']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=6).value = 'sin datos'

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=7).value = una_lista_json[json]['nPhone']['level4']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=7).value = 'sin datos'

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=8).value = una_lista_json[json]['nPhone']['level5']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=8).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=9).value = una_lista_json[json]['nPhone']['postalCode']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=9).value = 'sin datos'   
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=10).value = una_lista_json[json]['nPhone']['ddi']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=10).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=11).value = una_lista_json[json]['nPhone']['ddn']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=11).value = 'sin datos'               

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=12).value = una_lista_json[json]['nPhone']['characteristic']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=12).value = 'sin datos'

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=13).value = una_lista_json[json]['nPhone']['phoneNumber']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=13).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=14).value = una_lista_json[json]['nPhone']['validated']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=14).value = 'sin datos'  
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=15).value = una_lista_json[json]['nPhone']['additionalData']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=15).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=16).value = una_lista_json[json]['nPhone']['merlinCustomValues']['fullCellPhoneNumber']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=16).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=17).value = una_lista_json[json]['nPhone']['merlinCustomValues']['fullPhone']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=17).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=18).value = una_lista_json[json]['nPhone']['merlinCustomValues']['directory']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=18).value = 'sin datos'   
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=19).value = una_lista_json[json]['nPhone']['merlinCustomValues']['doNotCallRegistry']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=19).value = 'sin datos'
            
        una_hoja_grabando.cell(row=fila_escribiendo, column=20).value = 'PE'
        una_hoja_grabando.cell(row=fila_escribiendo, column=21).value = una_lista_id[json]['level2']
        una_hoja_grabando.cell(row=fila_escribiendo, column=22).value = una_lista_id[json]['level3']                   
        una_hoja_grabando.cell(row=fila_escribiendo, column=23).value = una_lista_id[json]['level4']            
        una_hoja_grabando.cell(row=fila_escribiendo, column=24).value = una_lista_id[json]['level5']
        una_hoja_grabando.cell(row=fila_escribiendo, column=25).value = una_lista_id[json]['phoneNumber']            
        una_hoja_grabando.cell(row=fila_escribiendo, column=26).value = una_lista_id[json]['characteristic']
        una_hoja_grabando.cell(row=fila_escribiendo, column=27).value = una_lista_id[json]['postalCode']           
        una_hoja_grabando.cell(row=fila_escribiendo, column=28).value = una_lista_id[json]['prefix']           
        una_hoja_grabando.cell(row=fila_escribiendo, column=29).value = una_lista_id[json]['additionalData']            
        una_hoja_grabando.cell(row=fila_escribiendo, column=30).value = str(una_lista_json[json])           
        fila_escribiendo = fila_escribiendo + 1
    print(hora_actual() + " - Se han volcado " + str(fila_escribiendo-2) + " registros en el archivo")
    
    