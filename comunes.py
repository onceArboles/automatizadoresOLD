import openpyxl
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os

BLACK = "\033[30m"
RED = "\033[31m"
GREEN = "\033[32m"
YELLOW = "\033[33m"
RED = "\033[34m"
MAGENTA = "\033[35m"
CYAN = "\033[36m"
WHITE = "\033[37m"
RESET = "\033[0m"
BOLD = "\033[1m"
UNDERLINE = "\033[4m"
BACKGROUND_RED = "\033[41m"
BACKGROUND_GREEN = "\033[42m"

def agregar_time_stamp(un_nombre_archivo: str) -> str:
    fecha_hora = str(datetime.datetime.now())
    fecha_hora = fecha_hora[0:19]
    fecha_hora = fecha_hora.replace(":","_")
    un_nombre_archivo = un_nombre_archivo + fecha_hora + ".xlsx"
    return un_nombre_archivo

def hora_actual() -> str:
    hora = datetime.datetime.now()
    hora = str(hora)
    hora = hora[11:19]
    return hora

def leer_singleLine(una_hoja, una_fila):
    diccionario = {}
    diccionario['id'] = una_hoja.cell(row = una_fila, column = 1).value
    diccionario['singleLine'] = una_hoja.cell(row = una_fila, column = 2).value
    return diccionario

def concatenarCoordenadas(lat: str, lon: str) -> str:
    lat = lat.replace(",", ".")
    lon = lon.replace(",",".")
    return (lat + ", " + lon)

def resumenInicio(una_libreria: str, una_cant_registros: int, un_pais:str, un_archivo:str) -> str:
    limpiar_pantalla()
    print(MAGENTA + "*** Procesar Lote por " + una_libreria + " ***" + RESET)
    print("Archivo a procesar: " + un_archivo)
    print("Cantidad de registros a procesar: " + str(una_cant_registros))
    print("Librería: "+ una_libreria)
    print("País: " + un_pais.upper())
    print("-----------------------------------------------------")
    seguir = str(input(YELLOW + "Ingrese 1 para continuar, de lo contrario presione cualquier otra tecla para salir --> " + RESET))
    if seguir == "1":
        hora = hora_actual()
        print(hora + " - Se comienza a procesar el lote. No cierres la consola durante el proceso")
        return hora
    else:
        exit()

def limpiar_pantalla():
    if os.name == "posix":
        os.system("clear")  # Para Unix/Linux/MacOS/BSD
    elif os.name in ["ce", "nt", "dos"]:
        os.system("cls")  # Para DOS/Windows

def mensaje_saludo():
    print("Gracias por usar el procesador de lote. Se termina la ejecución de programa. ")
    print("Merlin Data Quality 2024")

