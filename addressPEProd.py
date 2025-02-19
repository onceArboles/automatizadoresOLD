import openpyxl
import requests
import json
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os
from comunes import agregar_time_stamp
from comunes import hora_actual

from tokenSecurityProd import *

def procesarLoteAddressPEProd():
    #Variables comunes
    token = pedir_token_prod()
    print(token)
    libro = openpyxl.load_workbook('./Lotes a procesar/entrada_addressPE.xlsx')
    hoja = libro.active
    max_filas = hoja.max_row
    
    #en cabecera_archivo se definen los campos de la cabecera del archivo de salida
    cabecera_archivo = ['AddressPEProd_id','AddressPEProd_status', 'AddressPEProd_statusReason', 
                        'AddressPEProd_geoType', 'AddressPEProd_postalCode', 'AddressPEProd_street', 'AddressPEProd_houseNumber', 
                        'AddressPEProd_level1', 'AddressPEProd_level2', 'AddressPEProd_level3', 'AddressPEProd_level4', 
                        'AddressPEProd_level5', 'AddressPEProd_betweenStreet1','AddressPEProd_betweenStreet2', 
                        'AddressPEProd_unit', 'AddressPEProd_floor', 'AddressPEProd_additionalData', 'AddressPEProd_idSegment', 
                        'AddressPEProd_alternativeAddresses','AddressPEProd_latitude', 'AddressPEProd_longitude', 
                        'AddressPEProd_additionalPostalCode', 'AddressPEProd_fromStreetNumber', 'AddressPEProd_toStreetNumber', 
                        'AddressPEProd_placeReference', 'AddressPEProd_streetType', 'AddressPEProd_corner', 
                        'AddressPEProd_placeType', 'AddressPEProd_place', 'IN_level1', 'IN_level2', 'IN_level3','IN_level4','IN_level5',
                        'IN_postalCode', 'IN_street', 'IN_houseNumber', 'IN_floor', 'IN_unit', 'IN_additionalData', 'IN_fromStreetNumber', 
                        'IN_toStreetNumber', 'JsonOut']
    
    #En esta variable se almacenan los ID y las direcciones que se leen
    lista_id = [] 
    
    #En esta variable se almacenan cada uno de las respuestas de la API
    lista_json = [] 

    #--------------------------------------------
    
    resumenInicio("Address", max_filas-1, "PE")
    
    for fila in range (2, max_filas + 1):
        diccionario = leer_direccion_peru(hoja, fila) #leo una fila del archivo
        lista_id.append(diccionario) #agrego la fila leida a la lista de direcciones leidas        
        jsonSalida = llamar_AddressPe(diccionario,token) #llamo a Address
        lista_json.append(jsonSalida) #agrego la salida de Address a la lista de Json de salida
        print("Se han procesado " + str(fila-1) + " de " + str(max_filas-1) + " registros con Address Perú en Producción")  
    print(hora_actual() + " - Se han procesado correctamente con Address Perú en Producción " + str(len(lista_json)) + " registros.\nEspere a la generación del archivo de salida por favor")
    generar_archivo_addressPe('Lote_procesado_addressPEprod - ', cabecera_archivo, lista_json, lista_id) 


def llamar_AddressPe(un_diccionario, un_token):
    url = "https://ws.merlindataquality.com/address/pe/normalize"

    payload = json.dumps({
    'level1': 'PE',
    'level2': un_diccionario['level2'],
    "level3": un_diccionario['level3'],
    "level4": un_diccionario['level4'],
    "level5": un_diccionario['level5'],
    "postalCode": un_diccionario['postalCode'],
    "street": un_diccionario['street'],
    "houseNumber": un_diccionario['houseNumber'],
    "floor": un_diccionario['floor'],
    "unit": un_diccionario['unit'],
    "additionalData": un_diccionario['additionalData'],
    "toStreetNumber": un_diccionario['fromStreetNumber'],
    "fromSteetNumber": un_diccionario['toStreetNumber']})
    
    headers = {
    #'Authorization': "30910e0631407e6d29fc31e30fb66671",
    'Authorization': un_token,
    'Content-Type': 'application/json'}

    response = requests.request("POST", url, headers=headers, data=payload)
    respuesta_en_json = response.json()
    return respuesta_en_json
    
def leer_direccion_peru(una_hoja, una_fila):
    diccionario = {}
    diccionario['idRegister'] = una_hoja.cell(row = una_fila, column = 1).value
    if una_hoja.cell(row = una_fila, column = 2).value == None:
        diccionario['level2'] = ''
    else:
        diccionario['level2'] = una_hoja.cell(row = una_fila, column = 2).value


    if una_hoja.cell(row = una_fila, column = 3).value == None:
        diccionario['level3'] = ''
    else:
        diccionario['level3'] = una_hoja.cell(row = una_fila, column = 3).value
        
        
    if una_hoja.cell(row = una_fila, column = 4).value == None:
        diccionario['level4'] = ''
    else:
        diccionario['level4'] = una_hoja.cell(row = una_fila, column = 4).value  
        
        
    if una_hoja.cell(row = una_fila, column = 5).value == None:
        diccionario['level5'] = ''
    else:
        diccionario['level5'] = una_hoja.cell(row = una_fila, column = 5).value

    if una_hoja.cell(row = una_fila, column = 6).value == None:
        diccionario['postalCode'] = ''
    else:
        diccionario['postalCode'] = str(una_hoja.cell(row = una_fila, column = 6).value)

    if una_hoja.cell(row = una_fila, column = 7).value == None:
        diccionario['street'] = ''
    else:
        diccionario['street'] = una_hoja.cell(row = una_fila, column = 7).value

    if una_hoja.cell(row = una_fila, column = 8).value == None:
        diccionario['houseNumber'] = ''
    else:
        diccionario['houseNumber'] = una_hoja.cell(row = una_fila, column = 8).value

    if una_hoja.cell(row = una_fila, column = 9).value == None:
        diccionario['floor'] = ''
    else:
        diccionario['floor'] = una_hoja.cell(row = una_fila, column = 9).value

    if una_hoja.cell(row = una_fila, column = 10).value == None:
        diccionario['unit'] = ''
    else:
        diccionario['unit'] = una_hoja.cell(row = una_fila, column = 10).value
        
    if una_hoja.cell(row = una_fila, column = 11).value == None:
        diccionario['additionalData'] = ''
    else:
        diccionario['additionalData'] = una_hoja.cell(row = una_fila, column = 11).value

    if una_hoja.cell(row = una_fila, column = 12).value == None:
        diccionario['fromStreetNumber'] = ''
    else:
        diccionario['fromStreetNumber'] = una_hoja.cell(row = una_fila, column = 12).value

    if una_hoja.cell(row = una_fila, column = 13).value == None:
        diccionario['toStreetNumber'] = ''
    else:
        diccionario['toStreetNumber'] = una_hoja.cell(row = una_fila, column = 13).value
    return diccionario

def generar_archivo_addressPe(un_nombre, una_cabecera, una_lista_json, una_lista_id):
    #la ruta se debe expresar así: "./Documents/Python/Pepito2.xlsx"
    libroNuevo = Workbook()
    hoja = libroNuevo.active
    #Esto graba la cabecera del archivo
    print(hora_actual() + " - Se está creando el archivo '" + un_nombre + "' en el directorio local ")
    for campo in range (0, len(una_cabecera)):
        hoja.cell(row=1, column=campo+1).value = una_cabecera[campo]
    grabar_linea_addressPe(una_lista_json, una_lista_id, hoja)
    libroNuevo.save("./Lotes procesados/"+agregar_time_stamp(un_nombre))


def grabar_linea_addressPe(una_lista_json, una_lista_id, una_hoja_grabando):
    print(hora_actual() + " - Se comienza a volcar los datos en el archivo de salida")
    fila_escribiendo = 2
    for json in range (0, len(una_lista_id)):
        una_hoja_grabando.cell(row=fila_escribiendo, column=1).value = una_lista_id[json]['idRegister']
        
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=2).value = una_lista_json[json]['status']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=2).value = 'sin datos'

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=3).value = una_lista_json[json]['statusReason']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=3).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=4).value = una_lista_json[json]['nAddress']['geoType']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=4).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=5).value = una_lista_json[json]['nAddress']['postalCode']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=5).value = 'sin datos'            

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=6).value = una_lista_json[json]['nAddress']['street']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=6).value = 'sin datos'

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=7).value = una_lista_json[json]['nAddress']['houseNumber']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=7).value = 'sin datos'

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=8).value = una_lista_json[json]['nAddress']['level1']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=8).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=9).value = una_lista_json[json]['nAddress']['level2']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=9).value = 'sin datos'   
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=10).value = una_lista_json[json]['nAddress']['level3']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=10).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=11).value = una_lista_json[json]['nAddress']['level4']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=11).value = 'sin datos'               

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=12).value = una_lista_json[json]['nAddress']['level5']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=12).value = 'sin datos'

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=13).value = una_lista_json[json]['nAddress']['betweenStreet1']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=13).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=14).value = una_lista_json[json]['nAddress']['betweenStreet2']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=14).value = 'sin datos'  
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=15).value = una_lista_json[json]['nAddress']['unit']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=15).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=16).value = una_lista_json[json]['nAddress']['floor']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=16).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=17).value = una_lista_json[json]['nAddress']['additionalData']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=17).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=18).value = una_lista_json[json]['nAddress']['idSegment']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=18).value = 'sin datos'   
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=19).value = una_lista_json[json]['nAddress']['alternativeAddresses']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=19).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=20).value = una_lista_json[json]['nAddress']['latitude']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=20).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=21).value = una_lista_json[json]['nAddress']['longitude']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=21).value = 'sin datos'              
                    
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=22).value = una_lista_json[json]['nAddress']['additionalPostalCode']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=22).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=23).value = una_lista_json[json]['nAddress']['fromStreetNumber']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=23).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=24).value = una_lista_json[json]['nAddress']['toStreetNumber']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=24).value = 'sin datos'    

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=25).value = una_lista_json[json]['nAddress']['placeReference']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=25).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=26).value = una_lista_json[json]['nAddress']['streetType']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=26).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=27).value = una_lista_json[json]['nAddress']['corner']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=27).value = 'sin datos'    

        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=28).value = una_lista_json[json]['nAddress']['placeType']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=28).value = 'sin datos'
            
        try:
            una_hoja_grabando.cell(row=fila_escribiendo, column=29).value = una_lista_json[json]['nAddress']['place']
        except:
            una_hoja_grabando.cell(row=fila_escribiendo, column=29).value = 'sin datos'
        una_hoja_grabando.cell(row=fila_escribiendo, column=30).value = 'PE'
        una_hoja_grabando.cell(row=fila_escribiendo, column=31).value = una_lista_id[json]['level2']
        una_hoja_grabando.cell(row=fila_escribiendo, column=32).value = una_lista_id[json]['level3']                   
        una_hoja_grabando.cell(row=fila_escribiendo, column=33).value = una_lista_id[json]['level4']            
        una_hoja_grabando.cell(row=fila_escribiendo, column=34).value = una_lista_id[json]['level5']            
        una_hoja_grabando.cell(row=fila_escribiendo, column=35).value = una_lista_id[json]['postalCode']
        una_hoja_grabando.cell(row=fila_escribiendo, column=36).value = una_lista_id[json]['street']           
        una_hoja_grabando.cell(row=fila_escribiendo, column=37).value = una_lista_id[json]['houseNumber']           
        una_hoja_grabando.cell(row=fila_escribiendo, column=38).value = una_lista_id[json]['floor']            
        una_hoja_grabando.cell(row=fila_escribiendo, column=39).value = una_lista_id[json]['unit']             
        una_hoja_grabando.cell(row=fila_escribiendo, column=40).value = una_lista_id[json]['additionalData']             
        una_hoja_grabando.cell(row=fila_escribiendo, column=41).value = una_lista_id[json]['fromStreetNumber']
        una_hoja_grabando.cell(row=fila_escribiendo, column=42).value = una_lista_id[json]['toStreetNumber']
        una_hoja_grabando.cell(row=fila_escribiendo, column=43).value = str(una_lista_json[json])           
        fila_escribiendo = fila_escribiendo + 1
    print(hora_actual() + " - Se han volcado " + str(fila_escribiendo-2) + " registros en el archivo")  