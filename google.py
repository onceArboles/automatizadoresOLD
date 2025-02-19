import openpyxl
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os
from comunes import *
from geotypes import *

def procesar_lote_google(un_file): #esta es la función principal que llama a las otras
    #VARIABLES COMUNES
    libro = openpyxl.load_workbook(un_file)
    hoja = libro.active 
    libroParametros = openpyxl.load_workbook('parametros.xlsx')
    hojaParametros = libroParametros.active
    country = hojaParametros.cell(row=4, column=2).value
    max_filas = hoja.max_row
    max_columnas = hoja.max_column
    cabecera_archivo = ['google-IN-id','google-IN-original_single_line','google-partial_match','google-status','google-location_type',
                        'google-formatted_address','google-lat','google-lng','google-route_long',
                        'google-street_number_long','google-locality_long','google-neighborhood_long',
                        'google-administrative_area_level_2_long', 'google-administrative_area_level_1_long',
                        'google-country_long','google-postal_code_long','google-postal_code_suffix_long','google-types',
                        'google-CUSTOM-json_full','google-CUSTOM-geoType']
    lista_json = [] # En esta variable se almacenan cada uno de las respuestas de Google
    lista_id = [] # En esta lista se almacenan todos los ID y Single Line que se leyeron
    #----------------------------
    resumenInicio("Google", max_filas-1, country[8:10], "entrada_google.xlsx")

    print(hora_actual() + " - Se comienza a procesar el lote de " + str(max_filas-1) + " registros con Google. Espere por favor.")
    #este for itera leyendo una fila del archivo original, retornandola como un diccionario y enviandola a google.     
    for fila in range (2, max_filas + 1):
        diccionario = leer_singleLine(hoja, fila)
        lista_id.append(diccionario)
        jsonSalida = llamar_google(diccionario,country)
        lista_json.append(jsonSalida)
        print("Se han procesado " + str(fila-1) +" de " + str(max_filas-1) + " registros")
    print(hora_actual() + " - Se han procesado correctamente con Google " + str(len(lista_json)) + " registros.\nEspere a la generación del archivo de salida por favor")
    generar_archivo_google("OUT_Google ", cabecera_archivo, lista_json, lista_id) #se llama a la función para crear el archivo."""
    
def llamar_google(un_diccionario,component_country):
    endpoint = 'https://maps.googleapis.com/maps/api/geocode/json?'
    apiKey = 'AIzaSyD-JaQvEJXUEDbTmhL_vDkzzPju_NhBxN0'
    parametros = dict(key = apiKey,components = component_country, address=un_diccionario['singleLine'])
    respuesta = requests.get(url=endpoint, params=parametros)
    respuesta_en_json = respuesta.json()
    #print(respuesta_en_json)
    return respuesta_en_json

def generar_archivo_google(un_nombre, una_cabecera, una_lista_json, una_lista_id):
#la ruta se debe expresar así: "./Documents/Python/Pepito2.xlsx"
    libroNuevo = Workbook()
    hoja = libroNuevo.active
    hoja.title = "SalidaGoogle"
    #Esto graba la cabecera del archivo
    print(hora_actual() + " - Se está creando el archivo '" + un_nombre + "' en el directorio local ")    
    for campo in range (0, len(una_cabecera)):
        hoja.cell(row=1, column=campo+1).value = una_cabecera[campo]
    grabar_linea_google(una_lista_json, una_lista_id, hoja)
    libroNuevo.save("./Lotes procesados/"+agregar_time_stamp(un_nombre))       
    
def grabar_linea_google(una_lista_json, una_lista_id, una_hoja_grabando):
    print(hora_actual() + " - Se comienza a volcar los datos en el archivo de salida")
    #print(una_lista_json['results'])
    for fila in range (0, len(una_lista_id)):     
        una_hoja_grabando.cell(row=fila+2, column=1).value = una_lista_id[fila]['id'] #ID
        una_hoja_grabando.cell(row=fila+2, column=2).value = una_lista_id[fila]['singleLine'] #singleLine
        
        try:
            una_hoja_grabando.cell(row=fila+2, column=3).value = una_lista_json[fila]['status'] #status
        except:
            una_hoja_grabando.cell(row=fila+2, column=3).value = "Error"
            
        try:
            una_hoja_grabando.cell(row=fila+2, column=5).value = una_lista_json[fila]['results'][0]['geometry']['location_type'] #location_type
        except:
            una_hoja_grabando.cell(row=fila+2, column=5).value = "Error"
            
        try:
            una_hoja_grabando.cell(row=fila+2, column=6).value = una_lista_json[fila]['results'][0]['formatted_address'] #formatted_address        
        except:
            una_hoja_grabando.cell(row=fila+2, column=6).value = "Error"
            
        try:
            una_hoja_grabando.cell(row=fila+2, column=7).value = una_lista_json[fila]['results'][0]['geometry']['location']['lat'] #lon
        except:
            una_hoja_grabando.cell(row=fila+2, column=7).value = "Error"            
            
        try:
            una_hoja_grabando.cell(row=fila+2, column=8).value = una_lista_json[fila]['results'][0]['geometry']['location']['lng'] #lon
        except:
            una_hoja_grabando.cell(row=fila+2, column=8).value = "Error"
        
        route = None
        types = None
        streetNumber = None

        try:    
            for component in range (0, len(una_lista_json[fila]['results'][0]['address_components'])):
                if una_lista_json[fila]['results'][0]['address_components'][component]['types'][0] == 'route': #Route
                    una_hoja_grabando.cell(row=fila+2, column=9).value = una_lista_json[fila]['results'][0]['address_components'][component]['long_name']
                    route = str(una_lista_json[fila]['results'][0]['address_components'][component]['long_name'])
                    
                elif una_lista_json[fila]['results'][0]['address_components'][component]['types'][0] == 'street_number': #Street Number
                    una_hoja_grabando.cell(row=fila+2, column=10).value = una_lista_json[fila]['results'][0]['address_components'][component]['long_name']
                    streetNumber = str(una_lista_json[fila]['results'][0]['address_components'][component]['long_name'])
                    
                elif una_lista_json[fila]['results'][0]['address_components'][component]['types'][0] == 'locality': #Locality
                    una_hoja_grabando.cell(row=fila+2, column=11).value = una_lista_json[fila]['results'][0]['address_components'][component]['long_name']
                elif una_lista_json[fila]['results'][0]['address_components'][component]['types'][0] == 'neighborhood': #Neighborhood
                    una_hoja_grabando.cell(row=fila+2, column=12).value = una_lista_json[fila]['results'][0]['address_components'][component]['long_name']
                elif una_lista_json[fila]['results'][0]['address_components'][component]['types'][0] == 'administrative_area_level_2': #Administrative_area_level_2
                    una_hoja_grabando.cell(row=fila+2, column=13).value = una_lista_json[fila]['results'][0]['address_components'][component]['long_name']
                elif una_lista_json[fila]['results'][0]['address_components'][component]['types'][0] == 'administrative_area_level_1': #Administrative_area_level_1
                    una_hoja_grabando.cell(row=fila+2, column=14).value = una_lista_json[fila]['results'][0]['address_components'][component]['long_name']
                elif una_lista_json[fila]['results'][0]['address_components'][component]['types'][0] == 'country': #Country
                    una_hoja_grabando.cell(row=fila+2, column=15).value = una_lista_json[fila]['results'][0]['address_components'][component]['long_name']
                elif una_lista_json[fila]['results'][0]['address_components'][component]['types'][0] == 'postal_code': #Postal Code
                    una_hoja_grabando.cell(row=fila+2, column=16).value = una_lista_json[fila]['results'][0]['address_components'][component]['long_name']
                elif una_lista_json[fila]['results'][0]['address_components'][component]['types'][0] == 'postal_code_suffix': #Postal Code_Suffix
                    una_hoja_grabando.cell(row=fila+2, column=17).value = una_lista_json[fila]['results'][0]['address_components'][component]['long_name']
        except:
            print("Error")
        
        try:
            una_hoja_grabando.cell(row=fila+2, column=18).value = str(una_lista_json[fila]['results'][0]['types'])
            types = str(una_lista_json[fila]['results'][0]['types'])
            
        except:
            una_hoja_grabando.cell(row=fila+2, column=18).value = "Error"
            
        try:
            una_hoja_grabando.cell(row=fila+2, column=19).value = str(una_lista_json[fila]) #json Completo
        except:
            una_hoja_grabando.cell(row=fila+2, column=19).value = "Error"

#Asigno el geotype
        try:
            una_hoja_grabando.cell(row=fila+2, column=20).value = google_all(types, route, streetNumber)
        except:
            una_hoja_grabando.cell(row=fila+2, column=20).value = "Error"
        

        print(hora_actual() + " Registro " + str(fila+1) + " de " + str(len(una_lista_id)) + " procesado correctamente")         
    print(hora_actual() + " - Se han grabado " + str(len(una_lista_json)) + " registros procesados por Google. Fin del proceso")