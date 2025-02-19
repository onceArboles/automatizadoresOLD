import openpyxl
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os
from comunes import *

def pedir_token_prod() -> str:
    """La función lee un ID de usuario en la hoja de parámetros y devuelve un string correspondiente al token generado.
    Returns:
        str: token generado para el ambiente de PROD
    """
    libroParametros = openpyxl.load_workbook('parametros.xlsx')
    hojaParametros = libroParametros.active
    client_id = str(hojaParametros.cell(row=9, column=2).value)
        
    params = {
    'clientId': client_id}

    headers = {
    'Content-Type': 'application/json'}
    
    endpoint = 'https://ws.merlindataquality.com/security/generate-token?'
    
    tokenDevuelto = requests.get(url=endpoint, params=params, headers=headers)
    tokenJson = tokenDevuelto.json()
    token = tokenJson['token']
    return(token)