import openpyxl
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os
from comunes import *

def pedir_token() -> str:
    """La función lee un ID de usuario en la hoja de parámetros y devuelve un string correspondiente al token generado.

    Returns:
        str: token generado para el ambiente de QA
    """
    libroParametros = openpyxl.load_workbook('parametros.xlsx')
    hojaParametros = libroParametros.active
    client_id = hojaParametros.cell(row=5, column=2).value
    #print(client_id)
        
    params = {
    'clientId': client_id}

    headers = {
    'Content-Type': 'application/json'}
    
    endpoint = 'https://wsqa.merlindataquality.com/security/generate-token?'
    
    tokenDevuelto = requests.get(url=endpoint, params=params, headers=headers)
    tokenJson = tokenDevuelto.json()
    token = tokenJson['token']
    
    return(token)